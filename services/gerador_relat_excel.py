# -*- coding: utf-8 -*-
"""
===============================================================================
Projeto    : Weather Forecast
Arquivo    : gerador_relat_excel.py 
Autor      : Emerson A. Silva
Data       : Wed Jul 22 22:25:40 2026
Versão     : 1.0
Python     : Python 3.13.14 | packaged by Anaconda, Inc. 

Descrição:
        Salva as informações exibidas no dashboad em um relatorio em Excel ja pré 
 formatado, é utilizado o arquivo de template que esta em 
  /template/excel/ReportTemplate.xlsx
    
  Abre o arquivo de modelo, coloca as informações atuais do clima e manda um BytesIO  
  
  Cálculo da largura dos gráficos no excel, para isto fiz uma regra de 3 a partir de uma imagem que 
foi gerada no codigo, no codigo a largura da image era de 60, no excel a figura ficou com 1,59 cm

                                      1,59 ---- 60 
Largura desejada no excel em cm      20,00 ---- X          
                                       x = (20,00 * 60)/1,59  
                                       x = 440 (arredondei para cima)
                                            
Histórico:
       22/07/2026 - Inicio 
       23/07/2026 - Alterações para melhoria da performance na geração do relatório
       23/07/2026 - O quadro do clima agora vira uma imagem feita com as informações do info_clima
                    Insersão dos gráficos apresentados na tela no relatório
       24/07/2026 - Alterações no layout do relatório para melhorar a visualização dos gráficos
       24/07/2026 - Ajustes para que os gráficos que vão para o excel fiquem iguais aos apresentados na tela
===============================================================================
"""
from copy import deepcopy
from io import BytesIO
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.dataframe import dataframe_to_rows
from PIL import Image, ImageOps
import plotly.graph_objects as go
import pandas as pd
from services.gerador_de_imagens import base64_para_imagem, quadro_clima_base64

__PATH_MODELS__=  Path("templates/excel")
__MODEL_FILE__ = "ReportTemplate.xlsx"


def pil_para_imagem_excel(imagem: Image.Image, largura: int = 80, altura: int = 60,) -> ExcelImage:
    buffer = BytesIO()
    imagem.convert("RGBA").save(buffer,format="PNG")
    buffer.seek(0)
    imagem_excel = ExcelImage(buffer)
    imagem_excel.width = largura
    imagem_excel.height = altura

    # Mantém o buffer vivo enquanto a planilha é salva
    imagem_excel._buffer_referencia = buffer

    return imagem_excel

def plotly_para_imagem_excel(fig: go.Figure,
                             largura_excel: int = 700,
                             altura_excel: int = 390,
                             exibir_rotulos: bool = False,
                             unidade_rotulos: str ="°",
                             largura_borda: int = 1,
                             cor_borda: str = "#202020"
                            ) -> ExcelImage:

    # Cria uma cópia para não alterar o gráfico exibido no Streamlit
    fig_exportacao = deepcopy(fig)

    if exibir_rotulos:
        for trace in fig_exportacao.data:
            if not isinstance(trace, go.Scatter):
                continue
    
            if trace.name == "Máxima":
                posicao_texto = "top center"
    
            elif trace.name == "Mínima":
                posicao_texto = "bottom center"
    
            else:
                continue
    
            trace.update(mode="lines+markers+text",
                         texttemplate="%{y:.0f}" + f"{unidade_rotulos}",
                         textposition=posicao_texto,
                         textfont=dict(family="Arial", size=10, color="#000000"),
                        cliponaxis=False
                      )

    # O tamanho lógico da exportação será o mesmo tamanho usado no Excel.
    # scale = 2 aumenta somente a resolução.
    fig_exportacao.update_layout(width=largura_excel,
                                 height=altura_excel, 
                                 title=dict(x=0.05, 
                                            xanchor="left", 
                                            y=0.98, 
                                            yanchor="top", 
                                            font=dict(family = "Arial", size = 20, color = "#262730")
                                           ),
                                 font=dict(family = "Arial", size = 13),
                                 legend=dict(orientation="h",
                                             x=0.6,
                                             xanchor="left",
                                             y=1.2,
                                             yanchor="bottom",
                                             font=dict(family = "Arial", size = 13)
                                            ),
                                 margin=dict(l = 10, r = 5, t = 50, b = 60)
                                )
    fig_exportacao.update_xaxes(tickangle=0,
                                automargin=True,
                                tickfont=dict(family = "Arial", size = 12, color="#000000")
                               )
    fig_exportacao.update_yaxes(automargin=True,
                                tickfont=dict(family = "Arial", size = 12),
                                title_font=dict(family="Arial", size=13)
                                )
    imagem_bytes = fig_exportacao.to_image(format="png",
                                           width=largura_excel,
                                           height=altura_excel,
                                           scale=2
                                        )
    imagem_pil = Image.open(BytesIO(imagem_bytes)).convert("RGB")

    if largura_borda > 0:
        imagem_pil = ImageOps.expand(imagem_pil, border=largura_borda, fill=cor_borda)

    buffer_imagem = BytesIO()
    imagem_pil.save(buffer_imagem,format="PNG",quality=100)
    buffer_imagem.seek(0)

    imagem_excel = ExcelImage(buffer_imagem)

    # Mantém aproximadamente o tamanho visual solicitado,
    # mesmo que o PNG tenha sido produzido em 2x.
    proporcao = imagem_pil.height / imagem_pil.width

    imagem_excel.width = largura_excel
    imagem_excel.height = int(largura_excel * proporcao)

    # Mantém o BytesIO vivo até o workbook ser salvo
    imagem_excel._buffer_imagem = buffer_imagem

    return imagem_excel

#Cria uma nova planilha e descarrega o conteudo do da lista clima_json
def gerar_planilha_Dados_Brutos(workbook: Workbook, clima_json: dict):
    nova_planilha = workbook.create_sheet(title="Dados_Brutos_Previsao")
    
    df = pd.DataFrame(clima_json)
    
    # Célula onde o DataFrame começará
    linha_inicial = 1
    coluna_inicial = 1  
    
    for i, row in enumerate(dataframe_to_rows(df.drop(columns='icone'), index = False, header = True), start = linha_inicial):
        for j, valor in enumerate(row, start = coluna_inicial):
            nova_planilha.cell(row=i, column=j, value = valor)
    
    return

def preencher_relatorio_clima_Tempo_Agora(clima_json: dict, 
                                          info_user_local: dict,
                                          previsoes_dict: dict, 
                                          previsoes: pd.DataFrame,
                                          fonte_previsao: str, 
                                          mapa_imet1: Image.Image,
                                          mapa_imet2: Image.Image,
                                          graf_temp_maxmin: go.Figure,
                                          graf_umid_maxmim: go.Figure,
                                          graf_chuva: go.Figure) -> BytesIO:
  
    BASE_DIR = Path(__file__).parent.parent
    caminho_modelo = Path(BASE_DIR / __PATH_MODELS__  / __MODEL_FILE__)
    buffer_file = BytesIO()
    
    if caminho_modelo.exists():
        # Abre o arquivo modelo
        workbook = load_workbook(caminho_modelo)
        # Seleciona a planilha que vai trabalhar 
        planilha: Worksheet = workbook["Tempo Agora"]
        
        if clima_json is not None:
            # Coloca o Titulo
            planilha["A1"] = clima_json["local_clima"]
            planilha["J1"] = clima_json["local_clima"]
            planilha["B2"] = clima_json["data_por_extenso"]
            planilha["K2"] = clima_json["data_por_extenso"]
            
            planilha["G2"] = "Local de emissao:"
            planilha["S2"] = "Local de emissao:"
            
            loc_user_txt =  f"{info_user_local['local']}\n"
            loc_user_txt += f"{info_user_local['coordenadas']}\n"
            loc_user_txt += f"{info_user_local['origem_coordenadas']}"
                            
            planilha["H2"] = loc_user_txt
            planilha["T2"] = loc_user_txt
        
            info = clima_json['rodape_info']
            texto_info = f"{info['texto_info']}\n"
            texto_info += f"{info['versao']}"
            
           
        
            #Gera a imagem do Quadro do clima e coloca no relatório
            img_quadro_clima = quadro_clima_base64(clima_json)
            imagem_excel = pil_para_imagem_excel(img_quadro_clima, largura = 530, altura = 310)
            planilha.add_image(imagem_excel, "B3")
            planilha["K19"] = texto_info
            planilha["C3"] = f"Fonte: {clima_json['fonte_dados']}"
        #Coloca a imagem o mapa do  IMETT de precipitação no relatório
        # Adiciona uma borda preta de 2 pixels
        mapa_imet2 = ImageOps.expand(mapa_imet2,
                                     border=1,
                                     fill="#000000"   # cor da borda
                                    )
        imagem_excel = pil_para_imagem_excel(mapa_imet2, largura=310, altura = 310)
        planilha.add_image(imagem_excel, "H3")
        
      
        #Coloca os gráficos no relatório   
        #Acima explico como chegar nos valores de largura e altura 
        imagem_excel = plotly_para_imagem_excel(graf_chuva, largura_excel=760, altura_excel = 400)  #755
        planilha.add_image(imagem_excel, "K3")
        
        imagem_excel = plotly_para_imagem_excel(graf_temp_maxmin, largura_excel = 760, altura_excel = 400, 
                                                exibir_rotulos=True,
                                                unidade_rotulos="°C")
        planilha.add_image(imagem_excel, "K6")
        
        imagem_excel = plotly_para_imagem_excel(graf_umid_maxmim, largura_excel = 760, altura_excel = 400, 
                                                exibir_rotulos=True,
                                                unidade_rotulos="%")
        planilha.add_image(imagem_excel, "K13")
        planilha["F20"] = f"Fonte: {fonte_previsao}"
        planilha["S20"] = f"Fonte: {fonte_previsao}"
       
        #Monta a tabela de previsões
        row_excel = 5
        for indice, registro in previsoes.iterrows():
           img = base64_para_imagem(registro.get("dia_bola"))
           if img is not None:
               imagem_excel = pil_para_imagem_excel(img, largura=60, altura = 60)
               planilha.add_image(imagem_excel, f"B{row_excel}")
           else:
               planilha[f"B{row_excel}"] = "SEM ICONE"
            
           imagem_pil = base64_para_imagem(registro.get("icone")) #weather_icon(registro.get("icone"))
           if imagem_pil is not None:
               imagem_excel = pil_para_imagem_excel(imagem_pil, largura = 60, altura = 60)
               planilha.add_image(imagem_excel, f"C{row_excel}")
               
           else:
               planilha[f"C{row_excel}"] = "SEM ICONE"
           
           img = base64_para_imagem(registro.get("temp_max_min"))
           if img is not None:
               imagem_excel = pil_para_imagem_excel(img, largura=70, altura = 60)
               planilha.add_image(imagem_excel, f"D{row_excel}")
           else:
               planilha[f"D{row_excel}"] = "SEM ICONE"
           
           img = base64_para_imagem(registro.get("Umidade e chuva"))
           if img is not None:
               imagem_excel = pil_para_imagem_excel(img, largura=140, altura = 60)
               planilha.add_image(imagem_excel, f"E{row_excel}")
           else:
               planilha[f"E{row_excel}"] = "SEM ICONE"
                      
           planilha[f"F{row_excel}"] = registro.get('Descrição')
           
           row_excel += 1
           
           
        #Descarrega o clima_json em uma nova planilha chamada 
        gerar_planilha_Dados_Brutos(workbook, previsoes_dict)
           
        workbook.save(buffer_file)
        buffer_file.seek(0) #faz o "cursor" do arquivo voltar para o início
    
    return buffer_file